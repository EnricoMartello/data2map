'''
 # @ Author: Enrico Martello
 # @ Create Time: 2021.06.05 20:38
 # @ Modified time: 2022.01.10 11:50
 # @ Description: extracts data from the excel file and is responsible for the selection of data type (noise vs EOR).
 '''

import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import PySimpleGUI as sg


def read_measures(dataname):
    """Read the measured values from the excel file

    Args:
        dataname (string): path of the file to be read
    """
    wb = load_workbook(dataname, data_only=True)
    sh = wb['Confezionamento']
    print(len(wb.sheetnames))
    data = list(sh.values)

    print(data)

    # print(wb.sheetnames)
    # for i in range(20):
    #	print(sh.cell(row=3, column=i+1).value)


def select_sheet(dataname):
    """Selects the sheet of interest where data is stored

    Args:
        dataname (string): path of the file to be read

    Returns:
        list: list containg the measures
    """
    wb = load_workbook(dataname, data_only=True)
    event, values = sg.Window(
        'Scegli un foglio di lavoro',
        [[sg.Text('Scegli->'),
          sg.Listbox(wb.sheetnames, size=(20, 3), key='LB')],
         [sg.Button('Ok'), sg.Button('Cancel')]]).read(close=True)

    # if event == 'Ok':
    # 	sg.popup(f'Hai scelto {values["LB"][0]}')
    # else:
    # 	sg.popup_cancel('User aborted')

    sh = wb[str(values["LB"][0])]
    data = []
    for rowNumber in range(1, sh.max_row + 1):
        if sh.cell(row=rowNumber, column=1).value != None:
            data.append(sh.cell(row=rowNumber, column=1).value)

    return data


def select_datatype():
    """Selection of the data measured (either noise or radiation)

    Returns:
        boolean: False (to be read as 0) for noise and True (i.e. 1) for radiation
    """
    event, values = sg.Window(
        'Scegli il tipo di misurazione effettuata',
        [[sg.Text('Scegli->'),
          sg.Listbox(["Sorgenti luminose", "Fonti di rumore"], size=(20, 3), key='LB')],
         [sg.Button('Ok'), sg.Button('Cancel')]]).read(close=True)

    # if event == 'Ok':
    # 	sg.popup(f'Hai scelto {values["LB"][0]}')
    # else:
    # 	sg.popup_cancel('User aborted')

    return (f'{values["LB"][0]}' == "Sorgenti luminose")
